import openpyxl as op
from openpyxl.cell import get_column_letter

ROW_5 = '{0: >10}||{1: >10}|{2: >10}|{3: >10}|{4: >10}|{5: >10}|'
HEADER_SEPARATOR = '==================================================================='
SEPARATOR = '-------------------------------------------------------------------'

CHOICES = '''
    displayed above are rows {0} to {1}, columns {2} to {3},
    choose what to do next:
    (n): next_worksheet{4}
    (r): display next 5 columns to the right{5}
    (d): display next 5 rows downwards{6}
    (x): exit the reading of this file
    '''

LEFT_CHOICE = '\n(l): display last 5 columns to the left\n'
PREVIOUS_CHOICE = '\n(p): previous worksheet\n'
UP_CHOICE = '\n(u): display last 5 rows upwards\n'


def check_rows(row_start, row_end):
    row_start = row_start if row_start >= 1 else 1
    row_end = row_end if ((row_end - row_start) == 4) else row_start + 4  # ensure range of 5

    return row_start, row_end


def check_cols(col_start, col_end):
    col_start = col_start if col_start >= 1 else 1
    col_end = col_end if ((col_end - col_start) == 4) else col_start + 4  # ensure range of 5

    return col_start, col_end


def check_worksheet(wb, ws_num):
    if wb is None:  # ensure we have a valid workbook object
        print('Workbook not valid')
        wb = op.workbook.WorkBook()

    ws_num = ws_num if not ws_num < 0 else 0  # check range on ws_num
    ws_num = ws_num if ws_num < len(wb.worksheets) else len(wb.worksheets) - 1  # check range on ws_num

    return wb, ws_num


def display_menu(wb, ws_num, row_start=1, row_end=5, col_start=1, col_end=5):
    """Display the main menu of options"""

    wb, ws_num = check_worksheet(wb, ws_num)
    row_start, row_end = check_rows(row_start, row_end)
    col_start, col_end = check_cols(col_start, col_end)

    previous = PREVIOUS_CHOICE if ws_num > 0 else ""  # display choice only if valid
    left = LEFT_CHOICE if col_end > 5 else ""  # display choice only if valid
    up = UP_CHOICE if row_end > 5 else ""  # display choice only if valid

    print(CHOICES.format(row_start, row_end, col_start, col_end, previous, left, up))


def display_rows(wb, ws_num, row_start=1, row_end=5, col_start=1, col_end=5):
    '''display the specified rows/columns of the spreadsheet'''

    wb, ws_num = check_worksheet(wb, ws_num)
    row_start, row_end = check_rows(row_start, row_end)
    col_start, col_end = check_cols(col_start, col_end)
    row_nums = [i for i in range(row_start, row_end + 1)]  # enumerate range of 5 row numbers
    col_letters = [get_column_letter(i) for i in range(col_start, col_end + 1)]  # get 5 column letters

    print(ROW_5.format("", *col_letters))
    print(HEADER_SEPARATOR)

    ws = wb.worksheets[ws_num]

    for row in row_nums:  # truncates and prints up to row_end, col_end number of cell values
        value_list = [str(ws.cell(row=row, column=col).value)
                      if ws.cell(row=row, column=col).value is not None else ""
                      for col in range(col_start, col_end + 1)]
        short_values = [value if len(value) <= 10 else value[:7] + '...' for value in value_list]
        print(ROW_5.format(row, *short_values))
        print(SEPARATOR)
    display_menu(wb, ws_num, row_start, row_end, col_start, col_end)
